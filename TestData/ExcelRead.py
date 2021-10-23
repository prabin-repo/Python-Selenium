import openpyxl
book = openpyxl.load_workbook("C:\\Users\\prabibeh\\Documents\\testData.xlsx")
sheet = book.active
cell = sheet.cell(row=2,column=2)
print(cell.value)

row_size =sheet.max_row
col_size = sheet.max_column
print(row_size)
print(col_size)

for i in range(1, row_size+1):
    for j in range(1, col_size+1):
        print(sheet.cell(row=i,column=j).value)


